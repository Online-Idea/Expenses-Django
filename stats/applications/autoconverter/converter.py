import datetime
import ftplib
import io
import logging
import os
import random
import re
import xml.etree.ElementTree as ET
from collections import Counter
from ftplib import FTP
from functools import reduce
from pathlib import Path
from urllib.parse import urlparse

import demoji
# import xlsxwriter
import pandas as pd
import requests
from django.db.models import QuerySet
from openpyxl import Workbook
from pandas import DataFrame
from telebot.types import InputFile

from applications.ads.models import Ad
from applications.autoconverter.models import *
from applications.autoconverter.onllline_base import onllline_worker
from libs.services.email_sender import send_email
from libs.services.management.commands.bot import bot, break_message_to_parts
from libs.services.utils import get_models_verbose_names, pandas_numeric_to_object
from stats.settings import env


# Список Конфигураций: POST http://151.248.118.19/Api/Configurations/GetList
# Список Папок с фото: POST http://151.248.118.19/Api/Stock/GetClients

# Прогон шаблона (3 запроса)
# POST http://151.248.118.19/Api/Stock/StartProcess
# POST http://151.248.118.19/Api/Stock/GetProcessStep
# POST http://151.248.118.19/Api/Log/GetByProcessId

def get_converter_tasks(task_ids: list = None) -> list[ConverterTask] | QuerySet[ConverterTask]:
    """
    Собирает Задачи конвертера. Если есть task_ids то берёт только их, иначе все активные
    :param task_ids: список id модели ConverterTask
    :return: Django QuerySet с Задачами конвертера
    """
    if task_ids:
        tasks = ConverterTask.objects.filter(id__in=task_ids)
        return [tasks.get(id=task_id) for task_id in task_ids]  # Чтобы был тот же порядок что и в task_ids
    else:
        return ConverterTask.objects.filter(active=True)


def get_price(task):
    """
    Полный цикл для одного клиента.
    :param task: строка из таблицы Задачи конвертера
    """
    logging.info(f'Запускаю конвертер по задаче {task.name}')
    template = converter_template(task)
    if template.empty:
        logging.info(f'По клиенту {task.client.name} пустой шаблон')
        return
    process_id = converter_post(task)
    price = converter_process_result(process_id, template, task)
    logs = converter_logs(process_id)
    logs_xlsx = logs_to_xlsx(logs, template, task)
    import_result = onllline_worker(task)
    message = f'Логи конвертера:\n{logs}'
    if import_result:
        message += f'\n\nОтчет импорта базы:\n{import_result}'
    bot_messages(message, logs_xlsx, price, task)
    with open(logs_xlsx, 'rb') as file:
        file_content = file.read()
        save_on_ftp(logs_xlsx, file_content)
    os.remove(logs_xlsx)
    logging.info(f'Задача {task.name} - прайс готов')
    return


def get_price_without_converter(task):
    """
    Полный цикл для одного клиента но без конвертера. В качестве стока используется наш прайс
    :param task:
    :return:
    """
    logging.info(f'Запускаю задачу без конвертера {task.name}')
    template = converter_template(task)
    if template.empty:
        logging.info(f'По клиенту {task.client.name} пустой шаблон')
        return
    price = converter_process_result(None, template, task)
    import_result = onllline_worker(task)
    message = ''
    if import_result:
        message += f'\n\nОтчет импорта базы:\n{import_result}'
    bot_messages(message, None, price, task)
    print(f'Клиент {task.slug} - прайс готов')


def converter_template(task):
    """
    Из стока клиента делает шаблон для конвертера
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    # Сохраняю сток клиента, делаю по нему шаблон для конвертера
    slug = task.slug
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    stock_path = f'converter/{slug}/stocks/stock_{slug}_{file_date}'

    # Получаю сток
    if task.stock_source == 'Ссылка':
        response = requests.get(url=task.stock_url)
    elif task.stock_source == 'POST-запрос':
        data = {
            'login': task.stock_post_login,
            'password': task.stock_post_password,
        }
        response = requests.post(url=task.stock_post_host, data=data)
    else:
        raise ValueError('Источник стока должен быть "Ссылка" или "POST-запрос"')

    # Получаю тип файла стока
    content_type = response.headers['content-type']
    xml_types = ['text/xml', 'application/xml']
    if any(ele in content_type for ele in xml_types):
        stock_path += '.xml'
        content_type = 'xml'
    elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
        stock_path += '.xlsx'
        content_type = 'xlsx'
    elif 'text/csv' in content_type:
        stock_path += '.csv'
        content_type = 'csv'

    # Сохраняю сток на ftp
    os.makedirs(os.path.dirname(stock_path), mode=0o755, exist_ok=True)
    with open(stock_path, mode='wb') as file:
        file.write(response.content)
    # with open(stock_path, mode='w', encoding=task.stock_fields.encoding) as file:
    #     file.write(response.text)
    save_on_ftp(stock_path, response.content)

    # Путь шаблона
    template_path = f'converter/{slug}/templates/template_{slug}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    task.template = template_path
    task.save()

    # Разные функции в зависимости от типа файла стока
    if content_type == 'xml':
        template = template_xml(stock_path, template_path, task)
    elif content_type == 'xlsx':
        template = template_xlsx_or_csv(stock_path, 'xlsx', template_path, task)
    elif content_type == 'csv':
        template = template_xlsx_or_csv(stock_path, 'csv', template_path, task)
    else:
        logging.error('Неверный формат файла, должен быть xml или xlsx')
        return None

    # Убираю лишние пробелы
    template = template.map(lambda x: x.strip() if isinstance(x, str) else x)
    # template = template.applymap(lambda x: x.replace(' ', '') if isinstance(x, str) else x)

    with open(template_path, mode='rb') as file:
        file_content = file.read()
        save_on_ftp(template_path, file_content)
    os.remove(stock_path)

    return template


def template_xml(stock_path, template_path, task):
    """
    Шаблон из xml стока
    :param stock_path: путь к файлу стока
    :param template_path: путь к файлу шаблона
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    # XML tree
    tree = ET.parse(stock_path)
    root = tree.getroot()

    # XLSX шаблон
    workbook = Workbook()
    workbook.active.title = 'Шаблон'
    sheet = workbook.active

    # Заголовки шаблона
    template_col = StockFields.TEMPLATE_COL
    for k, col in template_col.items():
        sheet.cell(row=1, column=col[1] + 1, value=col[0])

    # Данные шаблона
    fields = task.stock_fields
    exception_col = ['modification_code', 'options_code', 'images', 'modification_explained', 'description', 'vin']
    for i, car in enumerate(root.iter(fields.car_tag)):
        if stock_xml_filter(car, task):
            # Обычные поля
            for field in fields._meta.fields:
                field_val = getattr(fields, field.name)
                # Если не пусто И поле в полях шаблона И поле НЕ в исключениях
                if field_val and field.name in template_col and field.name not in exception_col:
                    cell = car.findtext(field_val)
                    try:
                        if cell.isnumeric():
                            cell = int(cell)
                    except AttributeError:
                        pass
                    sheet.cell(row=i + 2, column=template_col[field.name][1] + 1, value=cell)

            # Поля-исключения
            if ',' in fields.modification_code:  # Код модификации
                # Разделяет по запятой в список если есть значение. Убирает запятую из данных стока
                mod = [car.findtext(f).replace(',', '') for f in fields.modification_code.split(', ') if
                       car.findtext(f)]
                sheet.cell(row=i + 2, column=template_col['modification_code'][1] + 1, value=' | '.join(mod))
            else:
                sheet.cell(row=i + 2, column=template_col['modification_code'][1] + 1,
                           value=car.findtext(fields.modification_code))

            if fields.options_code:
                options = multi_tags(fields.options_code, car, ' ')  # Опции
                sheet.cell(row=i + 2, column=template_col['options_code'][1] + 1, value=options)

            if fields.images:
                if 'авилон' in task.client.name.lower():
                    # Особая обработка фото от Авилона
                    images = avilon_photos(fields.images, car)
                else:
                    images = multi_tags(fields.images, car, ' ')  # Фото клиента
                sheet.cell(row=i + 2, column=template_col['images'][1] + 1, value=images)

            if ',' in fields.modification_explained:  # Расш. модификации
                mod = [car.findtext(f) for f in fields.modification_explained.split(', ') if car.findtext(f)]
                sheet.cell(row=i + 2, column=template_col['modification_explained'][1] + 1, value=' | '.join(mod))
            else:
                sheet.cell(row=i + 2, column=template_col['modification_explained'][1] + 1,
                           value=car.findtext(fields.modification_explained))

            if fields.description:
                if ',' in fields.description:
                    descr = [car.findtext(f) for f in fields.description.split(', ') if car.findtext(f)]
                    descr = [f.replace("\\n", "") for f in descr]
                    sheet.cell(row=i + 2, column=template_col['description'][1] + 1, value=' \n\n '.join(descr))
                else:
                    sheet.cell(row=i + 2, column=template_col['description'][1] + 1,
                               # value=car.findtext(fields.description))
                               value=multi_tags(fields.description, car, '\n'))

            # VIN
            vin = car.findtext(fields.vin)
            if task.fill_vin:
                # Если меньше 17 символов то добавляю в начало столько 'X' сколько не хватает до 17
                vin = f"{'X' * (17 - len(vin))}{vin}" if len(vin) < 17 else vin

            sheet.cell(row=i + 2, column=template_col['vin'][1] + 1, value=vin)

            # Для обработки прайса когда нужно смотреть по стоку. Добавляю столбец к шаблону
            # extras = ConverterExtraProcessing.objects.filter(converter_task=task, source='Сток')
            task_has_source_stock = Conditional.objects.filter(converter_extra_processing__converter_task=task,
                                                               source='Сток')
            if task_has_source_stock:
                # Беру Обработки прайса и Новые значения текущего таска
                extras = ConverterExtraProcessing.objects.filter(converter_task=task, active=True)
                extras_new_changes = ConverterExtraProcessingNewChanges.objects.filter(
                    converter_extra_processing__in=extras)
                for extra in extras:
                    conditionals = list(Conditional.objects.filter(converter_extra_processing=extra)
                                        .values('field'))
                    curr_extra_new_changes = extras_new_changes.filter(converter_extra_processing=extra)

                    for new_change in curr_extra_new_changes:
                        # Когда нужно брать значение из другого столбца
                        if new_change.new_value[:4] == '%col':
                            column_name_extra = re.findall(r'"(.*?)"', new_change.new_value)[0]
                            column_name_extra += '__stock'
                            conditionals.append({'field': column_name_extra})

                    for cond in conditionals:
                        column_name = cond['field']
                        if '__stock' in column_name:
                            value = multi_tags(column_name.replace('__stock', ''), car, ' ')
                        else:
                            value = multi_tags(column_name, car, ' ')

                        if column_name not in template_col:
                            max_column = len(template_col)
                            template_col[column_name] = (column_name, max_column)
                            sheet.cell(row=1, column=max_column + 1, value=column_name)

                        column_number = template_col[column_name][1] + 1

                        sheet.cell(row=i + 2, column=column_number, value=value)

    # Удаляю пустые строки
    for row in reversed(range(1, sheet.max_row + 1)):
        if all(cell.value is None for cell in sheet[row]):
            sheet.delete_rows(row)

    # Удаляю лишние пробелы
    for row in workbook.active.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.value = cell.value.rstrip()

    workbook.save(filename=template_path)

    # Чтобы pandas не менял колонки со значением true или false на bool, оставлял как str
    # - тогда обработки прайса (ConverterExtraProcessing) будут верно срабатывать
    template_df = pd.read_excel(template_path, decimal=',')
    bool_cols_as_str = {}
    for col in template_df.columns:
        unique_values = set(x.strip().lower() for x in template_df[col].dropna().unique().astype(str))
        if unique_values.issubset({'true', 'false'}):
            bool_cols_as_str[col] = 'str'

    template_df = pd.read_excel(template_path, decimal=',', dtype=bool_cols_as_str)

    return template_df


def avilon_photos(field: str, element: ET.Element) -> str:
    """
    Особая настройка для Авилона. Если брать их фото в том порядке как они идут в xml,
    то порядок фото получается случайным. А так как Авилон к каждому фото указывает угол,
    с которого они снимали, то можно по этому углу выставлять в красивый порядок фото
    :param field: имя тэга с фото
    :param element: элемент автомобиля
    :return: фото в одной строке, разделённые запятой
    """

    def get_by_view_type(tags, view_type):
        return [tag for tag in tags if tag.attrib['view_type'] == view_type]

    tags = element.findall(field)

    exterior = get_by_view_type(tags, 'exterior')
    # Первое фото немного боком влево
    for element in exterior:
        if element.attrib['angle'] in ['40', '320']:
            element.set('angle', '-40')
            break
    exterior.sort(key=lambda x: float(x.get('angle')))

    detail = get_by_view_type(tags, 'detail')

    interior = get_by_view_type(tags, 'interior')

    images = [image.text for image in exterior]
    images += [image.text for image in detail]
    images += [image.text for image in interior]
    return ', '.join(images)


def stock_xml_filter(car, task):
    """
    Проверяет автомобиль из xml стока по фильтрам из ConverterFilter
    :param car: node автомобиля из xml стока
    :param task: task (запись) из таблицы Задачи конвертера
    :return: True если фильтры пройдены
    """
    filters = ConverterFilter.objects.filter(converter_task=task, active=True)
    dict_filters, stock_fields, result = [], [], []

    # Перевожу фильтры в словарь вида: {'values': values, 'condition': condition, 'field': field}
    for f in filters:
        # Значения
        if '`' in f.value:
            values = [val.replace('`', '').strip() for val in f.value.split('`,')]
        else:
            values = [f.value]

        # Поля (теги)
        if '/' in f.field:  # Путь к детям
            parent, child = f.field.rsplit('/', 1)

            if '@' not in f.field:  # Дети
                stock_fields = [tag.text for tag in car.find(parent) if tag.tag == child]

            else:  # Атрибуты
                attribute_name = f.field.split('@')[1].split('=')[0]
                attribute_value = f.field.split('"')[1].replace('"', '')
                for tag in car.find(parent):
                    for tag_attribute in tag.attrib.items():
                        if tag_attribute == (attribute_name, attribute_value):
                            stock_fields = [tag.text]

        else:  # Просто один тег
            stock_fields = [car.findtext(f.field)]

        for field in stock_fields:
            dict_filters.append({
                'values': values,
                'condition': f.condition,
                'field': field,
            })

    for sf in dict_filters:
        for value in sf['values']:
            result.append(xml_filter_conditions(value, sf['condition'], sf['field']))

    # Если длина фильтров равна сумме результатов значит каждый фильтр вернул True, соответственно автомобиль подходит
    return len(dict_filters) == sum(result)


def xml_filter_conditions(value, condition, stock_field):
    """
    Условия фильтра
    :param value: значение для проверки
    :param condition: условие
    :param stock_field: значение из стока
    :return: True если условие выполнено
    """
    if condition == ConverterFilter.GREATER_THAN:
        if not stock_field:
            stock_field = 0
        return eval(f'{int(stock_field)} {condition} {int(value)}')
    elif condition == ConverterFilter.LESS_THAN:
        if not stock_field:
            stock_field = 0
        return eval(f'{int(stock_field)} {condition} {int(value)}')
    elif 'with' not in condition:
        return eval(f'"{value}" {condition} "{stock_field}"')
    elif condition == ConverterFilter.STARTS_WITH:
        return stock_field.startswith(value)
    elif condition == ConverterFilter.NOT_STARTS_WITH:
        return not (stock_field.startswith(value))
    elif condition == ConverterFilter.ENDS_WITH:
        return stock_field.endswith(value)
    elif condition == ConverterFilter.NOT_ENDS_WITH:
        return not (stock_field.endswith(value))


def template_xlsx_or_csv(stock_path, filetype, template_path, task):
    """
    Шаблон из xlsx или csv стока
    :param stock_path: путь к файлу стока
    :param filetype: тип файла: 'xlsx' или 'csv'
    :param template_path: путь к файлу шаблона
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    if filetype == 'xlsx':
        df_stock = pd.read_excel(stock_path, decimal=',')
    elif filetype == 'csv':
        encoding = task.stock_fields.encoding
        df_stock = pd.read_csv(stock_path, decimal=',', sep=';', header=0, encoding=encoding)
    else:
        return 'Неверный формат файла, должен быть xlsx или csv'

    # Чтобы pandas не менял колонки со значением true или false на bool, оставлял как str
    # - тогда обработки прайса (ConverterExtraProcessing) будут верно срабатывать
    bool_cols_as_str = {}
    for col in df_stock.columns:
        unique_values = set(x.strip().lower() for x in df_stock[col].dropna().unique().astype(str))
        if unique_values.issubset({'true', 'false'}):
            bool_cols_as_str[col] = 'str'

    if filetype == 'xlsx':
        df_stock = pd.read_excel(stock_path, decimal=',', dtype=bool_cols_as_str)
    elif filetype == 'csv':
        encoding = task.stock_fields.encoding
        df_stock = pd.read_csv(stock_path, decimal=',', sep=';', header=0, encoding=encoding, dtype=bool_cols_as_str)

    # Проверяю если сток это наш прайс. На случай если прайс готов.
    # Если первые 4 столбца совпадают с our_price_first_4_cols И Исходный VIN в столбцах
    our_price_first_4_cols = ['Марка', 'Модель', 'Комплектация', 'Авто.ру Комплектация']
    if list(df_stock.columns[:4]) == our_price_first_4_cols and 'Исходный VIN' in df_stock.columns:
        # df_stock.loc[df_stock['Фото'].str.contains('gallery'), 'Фото'] = ''
        df_stock.T.reset_index().T.to_excel(template_path, sheet_name='Шаблон', header=False, index=False)
        return df_stock

    df_stock = stock_xlsx_filter(df_stock, task)

    fields = StockFields.objects.filter(pk=task.stock_fields.id)
    fields = fields.values()[0]
    template_col = StockFields.TEMPLATE_COL

    # Для столбцов которые состоят из нескольких других
    for k, field in fields.copy().items():
        if ',' in str(field):
            columns_split = field.split(', ')
            df_stock[template_col[k][0]] = df_stock[columns_split[0]].str.cat(df_stock[columns_split[1:]]
                                                                              .astype(str), sep=' | ')
            del fields[k]

    # Для столбцов у которых одинаковый источник.
    # Например когда 'Код цвета' и 'Расш. цвета' это одинаковый столбец в стоке с названием 'Цвет'
    same_origin_columns_counter = Counter(fields.values())
    del same_origin_columns_counter[None]
    df_stock_copy = pd.DataFrame()
    # Добавляю каждый такой столбец по отдельности в dataframe
    for k, v in fields.copy().items():
        if same_origin_columns_counter[v] > 1:
            df_stock_copy[template_col[k][0]] = df_stock[v]
            del fields[k]
    # Меняю имена столбцов
    swapped_cols = {v: template_col[k][0] for k, v in fields.items() if k in template_col}
    # Объединяю
    combined_dfs = [df_stock_copy, df_stock.rename(columns=swapped_cols)]
    df_stock_copy = pd.concat(combined_dfs, axis=1)

    # Добавляю недостающие столбцы
    cur_cols = list(df_stock_copy.columns.values)
    for k, col in template_col.items():
        if col[0] not in cur_cols:
            df_stock_copy[col[0]] = ''

    # Для обработки прайса когда нужно смотреть по стоку. Добавляю столбец к шаблону
    # extras = ConverterExtraProcessing.objects.filter(converter_task=task, source='Сток')
    task_has_source_stock = Conditional.objects.filter(converter_extra_processing__converter_task=task, source='Сток')
    if task_has_source_stock:
        # Беру Обработки прайса и Новые значения текущего таска
        extras = ConverterExtraProcessing.objects.filter(converter_extra_processing__converter_task=task, active=True)
        for extra in extras:
            conditionals = list(Conditional.objects.filter(converter_extra_processing=extra).values('field'))
            for cond in conditionals:
                column_name = cond.field
                if column_name not in template_col:
                    max_column = len(template_col)
                    template_col[column_name] = (column_name, max_column)

    # Оставляю только нужные столбцы в том порядке как в template_col
    df_stock_copy = df_stock_copy[[v[0] for k, v in template_col.items()]]
    # В Опции и пакеты заменяю переносы строк на пробел
    df_stock_copy['Опции и пакеты'].replace(r'\n', ' ', regex=True, inplace=True)
    # Убираю лишние пробелы
    df_stock_copy = df_stock_copy.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    df_stock_copy.T.reset_index().T.to_excel(template_path, sheet_name='Шаблон', header=False, index=False)

    return df_stock_copy


def stock_xlsx_filter(df, task):
    """
    Проверяет автомобиль из xlsx стока по фильтрам из ConverterFilter
    :param df: xlsx сток в виде pandas dataframe
    :param task: task (запись) из таблицы Задачи конвертера
    :return: Отфильтрованный dataframe
    """
    filters = ConverterFilter.objects.filter(converter_task=task, active=True)
    filter_strings = []
    for f in filters:
        filter_or = []
        if '`' in f.value:
            values = [val.replace('`', '').strip() for val in f.value.split('`,')]
        else:
            values = [f.value]

        for value in values:
            if f.condition == ConverterFilter.CONTAINS:
                filter_or.append(f'(df["{f.field}"].str.contains({value})')
            elif f.condition == ConverterFilter.NOT_CONTAINS:
                filter_or.append(f'(~df["{f.field}"].str.contains({value})')
            elif f.condition == ConverterFilter.EQUALS:
                filter_or.append(f'(df["{f.field}"] == "{value}")')
            elif f.condition == ConverterFilter.NOT_EQUALS:
                filter_or.append(f'(~df["{f.field}"] == "{value}")')
            elif f.condition == ConverterFilter.GREATER_THAN:
                filter_or.append(f'(df["{f.field}"] > "{value}")')
            elif f.condition == ConverterFilter.LESS_THAN:
                filter_or.append(f'(df["{f.field}"] < "{value}")')
            elif f.condition == ConverterFilter.STARTS_WITH:
                filter_or.append(f'(df["{f.field}"].str.startswith("{value}"))')
            elif f.condition == ConverterFilter.NOT_STARTS_WITH:
                filter_or.append(f'~(df["{f.field}"].str.startswith("{value}"))')
            elif f.condition == ConverterFilter.ENDS_WITH:
                filter_or.append(f'(df["{f.field}"].str.endswith("{value}"))')
            elif f.condition == ConverterFilter.NOT_ENDS_WITH:
                filter_or.append(f'~(df["{f.field}"].str.endswith("{value}"))')

        filter_strings.append(f'({" | ".join(filter_or)})')

    return eval(f'df.loc[{" & ".join(filter_strings)}]') if filter_strings else df


def price_extra_processing(df: DataFrame, task: ConverterTask, template: DataFrame) -> DataFrame:
    """
    Меняет данные в прайсе по условию
    :param df: прайс в виде pandas dataframe
    :param task: task (запись) из таблицы Задачи конвертера
    :param template: шаблон в виде pandas dataframe
    :return: Готовый прайс с изменениями
    """
    extra_processings = ConverterExtraProcessing.objects.filter(converter_task=task, active=True)
    conditionals = Conditional.objects.filter(converter_extra_processing__in=extra_processings)
    new_changes = ConverterExtraProcessingNewChanges.objects.filter(converter_extra_processing__in=extra_processings)

    price_columns = get_models_verbose_names(Ad)
    price_columns.append('Фото')

    # Если изменение зависит от данных в стоке то сначала эти данные добавляются к шаблону.
    # Потом шаблон передаётся сюда и добавляется к прайсу.
    if conditionals.filter(source='Сток'):
        template = template.add_suffix('_template')
        df = pd.merge(df, template, left_index=True, right_index=True)

    # Для каждого изменения по клиенту
    for change in new_changes:
        curr_conditionals = conditionals.filter(converter_extra_processing=change.converter_extra_processing)
        masks = []

        # Для каждого условия в изменении
        for cond in curr_conditionals:
            if cond.source == 'Сток':
                cond.field += '_template'

            # Если несколько значений
            if '`' in cond.value:
                values = [val.replace('`', '').strip() for val in cond.value.split('`,')]
            else:
                values = [cond.value]

            # Для каждого значения
            or_masks = []
            for value in values:
                if value.isdigit():
                    value = int(value)

                if cond.condition == ConverterFilter.CONTAINS:
                    or_masks.append(df[cond.field].str.contains(value))
                elif cond.condition == ConverterFilter.NOT_CONTAINS:
                    or_masks.append(~df[cond.field].str.contains(value))
                elif cond.condition == ConverterFilter.EQUALS:
                    or_masks.append(df[cond.field] == value)
                elif cond.condition == ConverterFilter.NOT_EQUALS:
                    or_masks.append(df[cond.field] != value)
                elif cond.condition == ConverterFilter.GREATER_THAN:
                    df[cond.field] = pd.to_numeric(df[cond.field], errors='coerce').dropna().astype(int)
                    or_masks.append(df[cond.field] > value)
                    df[cond.field] = df[cond.field].fillna('')
                elif cond.condition == ConverterFilter.GREATER_THAN_EQUALS:
                    df[cond.field] = pd.to_numeric(df[cond.field], errors='coerce').dropna().astype(int)
                    or_masks.append(df[cond.field] >= value)
                    df[cond.field] = df[cond.field].fillna('')
                elif cond.condition == ConverterFilter.LESS_THAN:
                    df[cond.field] = pd.to_numeric(df[cond.field], errors='coerce').dropna().astype(int)
                    or_masks.append(df[cond.field] < value)
                    df[cond.field] = df[cond.field].fillna('')
                elif cond.condition == ConverterFilter.LESS_THAN_EQUALS:
                    df[cond.field] = pd.to_numeric(df[cond.field], errors='coerce').dropna().astype(int)
                    or_masks.append(df[cond.field] <= value)
                    df[cond.field] = df[cond.field].fillna('')
                elif cond.condition == ConverterFilter.STARTS_WITH:
                    or_masks.append(df[cond.field].str.startswith(str(value)))
                elif cond.condition == ConverterFilter.NOT_STARTS_WITH:
                    or_masks.append(~df[cond.field].str.startswith(str(value)))
                elif cond.condition == ConverterFilter.ENDS_WITH:
                    or_masks.append(df[cond.field].str.endswith(str(value)))
                elif cond.condition == ConverterFilter.NOT_ENDS_WITH:
                    or_masks.append(~df[cond.field].str.endswith(str(value)))

            # Несколько значений через И если условие было отрицательным
            if cond.condition in ConverterFilter.NEGATIVE_CONDITIONS:
                combined_or_mask = reduce(lambda x, y: x & y, or_masks)
            # Иначе через ИЛИ
            else:
                combined_or_mask = reduce(lambda x, y: x | y, or_masks)
            combined_or_mask = combined_or_mask.fillna(False)
            masks.append(combined_or_mask)

        # Объединяю маски
        combined_mask = reduce(lambda x, y: x & y, masks)

        # Проставляю новые значения
        # Когда нужно брать значение из другого столбца
        if change.new_value and change.new_value[:4] == '%col':
            column = re.findall(r'"(.*?)"', change.new_value)[0]
            # if change.source == 'Сток':
            #     column += '__stock_template'
            if column not in price_columns:
                column += '__stock_template'
            df.loc[combined_mask, change.price_column_to_change] = df[column]
        # Когда нужно добавить в начало
        elif change.change_type == 'Добавить в начало':
            df.loc[combined_mask, change.price_column_to_change] = df.loc[combined_mask, change.price_column_to_change] \
                .apply(lambda x: change.new_value + str(x))
        # Когда нужно добавить в конец
        elif change.change_type == 'Добавить в конец':
            df.loc[combined_mask, change.price_column_to_change] = df.loc[combined_mask, change.price_column_to_change] \
                .apply(lambda x: str(x) + change.new_value)
        # Когда нужно заменить полностью
        elif change.change_type == 'Полностью':
            if df[change.price_column_to_change].dtype == 'float64':
                df[change.price_column_to_change] = df[change.price_column_to_change].astype(object)
            df.loc[combined_mask, change.price_column_to_change] = change.new_value

    df = df.drop(df.filter(regex='_template').columns, axis=1)

    return df


def multi_tags(field, element, delimiter):
    """
    Обрабатывает поля для которых данные собираются из нескольких тегов
    :param field: поле
    :param element: элемент из xml
    :param delimiter: разделитель
    :return: готовые данные для шаблона
    """
    result = []

    if '/' in field:
        parent = field.split('/')[0]
        parent_element = element.find(parent)
        if parent_element:
            if '@' not in field:  # Если тег с детьми и нужно значение детей
                result = [tag.text for tag in parent_element]
            else:  # Если тег с детьми и из детей нужен атрибут
                attribute = field.split('@')[1]
                result = [tag.attrib[attribute] for tag in parent_element]
    else:
        if '@' not in field:  # Если тег несколько раз повторяется и нужно значение
            tags = element.findall(field)
            if tags:
                result = [tag.text for tag in tags]
        else:  # Если тег несколько раз повторяется и из него нужен атрибут
            tag_name, attribute = field.split('@')
            tags = element.findall(tag_name)
            if tags:
                result = [tag.attrib[attribute] for tag in tags]

    return delimiter.join(result)


def converter_post(task):
    """
    Отправляет шаблон и настройки по которым нужно прогнать конвертер
    :param task: task (запись) из таблицы Задачи конвертера
    :return: processId по которому дальше можно получить прайс
    """
    # Отправляю шаблон
    url = 'http://151.248.118.19/Api/Stock/StartProcess'

    configuration = task.configuration.configuration if task.configuration is not None else Configuration.DEFAULT
    payload = {
        'client': task.photos_folder.folder,
        'configuration': configuration,
        'frontLength': task.front,
        'backLength': task.back,
        'interiorsLength': task.interior,
        'onlySalon': task.salon_only,
    }
    template = open(task.template, 'rb')
    files = {'file': ('template.xlsx', template, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      {'Expires': '0'})}
    response = requests.post(url=url, data=payload, files=files)
    template.close()
    os.remove(task.template)
    return response.json()['processId']


def converter_get_price_by_process_id(task, process_id):
    """
    # Получаю прайс от конвертера по process_id
    :param task: строка из таблицы Задачи конвертера
    :param process_id: из converter_post
    :return: прайс как pandas dataframe
    """
    client = task.slug
    url = 'http://151.248.118.19/Api/Stock/GetProcessResult'
    payload = {'processId': (None, process_id)}
    response = requests.post(url=url, files=payload)

    # Сохраняю прайс в xlsx
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    save_path_date = f'converter/{client}/prices/price_{client}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(save_path_date), exist_ok=True)
    with open(save_path_date, 'wb') as file:
        file.write(response.content)
    save_on_ftp(save_path_date, response.content)

    # Обработки прайса
    read_file = pd.read_excel(save_path_date, decimal=',')

    # Добавляю столбцы базы которых нет в прайсе после конвертера
    read_file.insert(37, 'Дата размещения', '')
    read_file.insert(38, 'Кол. в ПТС', '')
    read_file = read_file.drop(columns=['Блокировано', 'Блокировка на авто.ру'])
    end_columns = ['ComID', 'Онлайн-показ', 'Avito gen', 'Avito mod', 'Avito com', 'Avito mid',
                   'Гарантия производителя', 'Валюта', 'Авито аукцион', 'Двери', 'Гарантия на ремонт',
                   'КАСКО в подарок', 'Комплект шин в подарок', 'Скидка на КАСКО', 'ТО в подарок',
                   'Аксессуары в подарок', 'Ускоренное оформление', 'Спецусловия кредита']
    for col in end_columns:
        read_file[col] = ''

    os.remove(save_path_date)
    return read_file


def converter_process_result(process_id, template, task):
    """
    Возвращает готовый прайс
    :param process_id: из converter_post
    :param template: шаблон как pandas dataframe - если из шаблона нужны данные для прайса
    :param task: строка из таблицы Задачи конвертера
    """
    client = task.slug
    if task.use_converter:
        read_file = converter_get_price_by_process_id(task, process_id)
    else:
        read_file = template

    # Переношу Описание из шаблона если оно есть в шаблоне
    if 'Описание' in template.columns:
        read_file['Описание2'] = template['Описание']
        read_file.loc[read_file['Описание2'].notnull(), 'Описание'] = read_file['Описание2']
        read_file.drop(columns='Описание2', inplace=True)

    # Убираю автомобили которые не расшифрованы (пустые столбцы Марка, Цвет либо Фото)
    read_file = read_file[(~read_file['Марка'].isnull()) &
                          (~read_file['Цвет'].isnull()) &
                          (~read_file['Фото'].isnull())]
    # Если Максималка пустая то считать её как сумму других скидок
    read_file.loc[read_file['Максималка'].isna(), 'Максималка'] = read_file[['Трейд-ин', 'Кредит', 'Страховка']].sum(
        axis=1)

    read_file.fillna('', inplace=True)

    # Различные изменения прайса по условию
    read_file = price_extra_processing(read_file, task, template)

    # Мелкие замены
    read_file = pandas_numeric_to_object(read_file)
    read_file = read_file.astype(str).replace(
        {
            r"\.0$": "",
            "é": "e",
            "\u2070": "0",
            "\xb3": "",
            "\uff08": "",
            "\uff09": "",
        },
        regex=True,
    )
    read_file = read_file.map(lambda x: demoji.replace(x, ''))
    read_file['Описание'] = read_file['Описание'].replace('_x000d_', '', regex=True)

    # Добавляю объявления, которых нет в стоке клиента, с другого файла
    if task.add_to_price:
        if task.add_to_price.endswith('.csv'):
            add_manually_df = pd.read_csv(task.add_to_price, decimal=',', sep=';', header=0, encoding='cp1251')
        elif task.add_to_price.endswith('.xlsx'):
            add_manually_df = pd.read_excel(task.add_to_price, decimal=',')
        else:
            raise ValueError('Файл с добавлением объявлений должен быть csv или xlsx')
        float_columns = ['Цена', 'Трейд-ин', 'Кредит', 'Страховка', 'Максималка']
        add_manually_df[float_columns] = add_manually_df[float_columns].fillna(0).astype(int)
        read_file = pd.concat([read_file, add_manually_df], axis=0)

    if task.change_vin:
        read_file = change_vin_to_random(read_file)

    # Сохраняю в csv
    save_path = f'converter/{client}/prices/price_{client}.csv'

    # string_buffer = io.StringIO()
    # read_file.to_csv(string_buffer, sep=';', header=True, index=False, decimal=',')
    # csv_string = string_buffer.getvalue()
    #
    # with open(save_path, 'w', encoding='cp1251', errors='ignore') as f:
    #     f.write(csv_string)
    read_file.to_csv(save_path, sep=';', header=True, index=False, decimal=',', encoding='cp1251', errors='ignore',
                     lineterminator='\n')
    with open(save_path, 'rb') as file:
        file_content = file.read()
        save_on_ftp(save_path, file_content)

    task.price = save_path
    task.save()

    os.remove(save_path)
    return read_file


def converter_logs(process_id):
    """
    Логи конвертера
    :return: логи в виде текста
    """
    # Логи которые присылает конвертер
    url = 'http://151.248.118.19/Api/Log/GetByProcessId'
    payload = {'processId': process_id}
    response = requests.post(url=url, json=payload)
    logs = response.json()['log']
    logs = logs.replace(' ,', '')  # Убираю лишние запятые
    return logs


def logs_to_xlsx(logs, template, task):
    """
    Логи конвертера вместе с расшифровкой в xlsx
    :param logs: логи от converter_logs
    :param template: шаблон от converter_template
    :param task: строка из таблицы Задачи конвертера
    :return: xlsx файл логов
    """
    lookup_cols = {
        # База из лога: (Имя столбца с кодом, Имя столбца с расшифровкой)
        'Модификации': ('Код модификации', 'Расш. модификации'),
        'Цвета': ('Код цвета', 'Расш. цвета'),
        'Интерьеры': ('Код интерьера', 'Расш. интерьера'),
        'Комплектации': ('Код модификации', 'Расш. модификации'),
        # 'Опции': в логи идут только коды, без расшифровки
        # 'Фото': только количество без фото
    }
    client_name = task.client.name
    client_slug = task.slug

    # Переделываю логи в словарь
    lines = logs.split('\n')[:-2]  # Последние 2 убираю т.к. там Время обработки и пустая строка
    logs_dict = {}
    for line in lines:
        try:
            key = line.split('"')[1]
        except IndexError:
            continue
        start = line.index(':') + 1
        end = line.index(';')
        value = []
        for v in line[start:end].split(','):
            v = v.strip()
            if v.isnumeric():
                v = int(v)
            value.append(v)

        if key in ['Опции', 'Фото']:  # Опции без расшифровки, Фото только количество
            logs_dict[key] = pd.Series(value, name=key)
        else:  # Остальные это pandas dataframe в виде Кода и Расшифровки
            df2 = pd.Series(value, name='Код', dtype='string')
            joined = pd.merge(template, df2, left_on=lookup_cols[key][0], right_on='Код')
            joined.drop_duplicates(subset=[lookup_cols[key][0]], inplace=True)
            joined = joined[[lookup_cols[key][0], lookup_cols[key][1]]]
            logs_dict[key] = joined

    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    logs_save_path = f'converter/{client_slug}/logs/log_{client_slug}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(logs_save_path), exist_ok=True)

    # Готовые логи в xlsx
    with pd.ExcelWriter(logs_save_path) as writer:
        for key, value in logs_dict.items():
            df = pd.DataFrame(value)
            # Такой длинный вариант чтобы убрать форматирование заголовков которое pandas применяет по умолчанию
            df.T.reset_index().T.to_excel(writer, sheet_name=key, header=False, index=False)

    # Отправляю логи на почту
    if task.notifications_email:
        if any(code in logs_dict for code in list(lookup_cols.keys())) or logs_dict['Фото'].iloc[0] > 0:
            send_email(
                subject='Конвертер. Добавить коды',
                body=f'Нужно добавить коды для клиента {client_name}.\nЛоги:\n{logs}',
                recipients=task.notifications_email,
                attachments=[logs_save_path],
            )

    return logs_save_path


def bot_messages(logs, logs_xlsx, price, task):
    """
    Сообщения для телеграм бота
    :param logs: логи от converter_logs
    :param logs_xlsx: логи в xlsx от logs_to_xlsx
    :param price: прайс от converter_process_result
    :param task: строка из таблицы Задачи конвертера
    """
    client_slug = task.slug

    # Прайс в csv
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    price_save_path = f'converter/{client_slug}/prices/price_{client_slug}_{file_date}.csv'

    # string_buffer = io.StringIO()
    # price.to_csv(string_buffer, sep=';', header=True, index=False, decimal=',')
    # csv_string = string_buffer.getvalue()
    #
    # with open(price_save_path, 'w', encoding='cp1251', errors='ignore') as f:
    #     f.write(csv_string)
    price.to_csv(price_save_path, sep=';', header=True, index=False, decimal=',', encoding='cp1251', errors='ignore',
                 lineterminator='\n')

    # Отправка логов и прайса через бота телеграма
    chat_ids = ConverterLogsBotData.objects.all()
    logs = f'🟢 {task.name}\n\n{logs}'
    for chat_id in chat_ids:
        split_message = break_message_to_parts(logs)
        for message in split_message:
            bot.send_message(chat_id.chat_id, message)
        # if len(logs) > 4095:  # У телеграма ограничение на 4096 символов в сообщении
        #     for x in range(0, len(logs), 4095):
        #         bot.send_message(chat_id.chat_id, logs[x:x + 4095])
        # else:
        #     bot.send_message(chat_id.chat_id, logs)
        if logs_xlsx:
            bot.send_document(chat_id.chat_id, InputFile(logs_xlsx))
        bot.send_document(chat_id.chat_id, InputFile(price_save_path))

    os.remove(price_save_path)
    return


def change_vin_to_random(price):
    def modify_vin(vin):
        if len(vin) == 17:
            last_six = vin[-6:]
            if last_six.isdigit():
                # All last 6 characters are digits
                random_number = str(random.randint(0, 999999)).zfill(6)
                return vin[:-6] + random_number
            else:
                # Last 6 characters contain letters
                for i in range(5, -1, -1):
                    if last_six[i].isalpha():
                        # Found the last letter
                        prefix = last_six[:i + 1]
                        digits_part = last_six[i + 1:]
                        if digits_part.isdigit():
                            random_number = str(random.randint(0, int(digits_part))).zfill(len(digits_part))
                            return vin[:-6] + prefix + random_number
                        break
        return vin

    price['VIN'] = price['Исходный VIN'].apply(modify_vin)
    return price


def save_on_ftp(save_path, file):
    """
    Сохраняет файл на ftp
    :param save_path: путь к файлу
    :param file: binary файл
    """
    file_path = Path(save_path)
    # with FTP('ph.onllline.ru', env('FTP_LOGIN'), env('FTP_PASSWORD')) as ftp, open(save_path, 'rb') as file:
    with FTP('ph.onllline.ru', env('FTP_LOGIN'), env('FTP_PASSWORD')) as ftp:
        # TODO передавать несколько файлов, не только один
        cd_tree(ftp, str(file_path.parents[0]))
        # file = io.BytesIO(file)
        # ftp.storbinary(f'STOR {file_path.name}', file)
        ftp.storbinary(f'STOR {file_path.name}', io.BytesIO(file))
        ftp.cwd('/')
    return


def cd_tree(ftp: FTP, url_path: str) -> None:
    """
    Перемещается по папкам ftp по пути из url_path.
    Создаёт папки на ftp если они не существуют
    :param ftp: FTP класс из ftplib
    :param url_path: путь который нужен на ftp
    """
    url_path = url_path.replace('\\', '/')  # Костыль для Windows

    # Вытаскиваю путь из ссылки
    parsed_url = urlparse(url_path)
    path = parsed_url.path.strip('/')

    # Если ссылка не полная
    if not path.startswith('/'):
        path = '/' + path

    # Перемещаюсь по папкам из пути
    for folder in path.split('/'):
        try:
            ftp.cwd(folder)
        except ftplib.error_perm:  # Создаю папку если не существует
            ftp.mkd(folder)
            ftp.cwd(folder)
    return


def get_photo_folders():
    """ Получает от конвертера список папок с фото и добавляет новые в базу """
    url = 'http://151.248.118.19/Api/Stock/GetClients'
    response = requests.post(url).json()
    current_folders = [f.folder for f in PhotoFolder.objects.all()]
    new_folders = []
    for folder in response:
        if folder not in current_folders:
            new_folders.append(PhotoFolder(folder=folder))
    PhotoFolder.objects.bulk_create(new_folders)
    return


def get_configurations():
    """ Получает от конвертера конфигурации и добавляет/обновляет в базе """
    url = 'http://151.248.118.19/Api/Configurations/GetList'
    response = requests.post(url).json()
    current_configurations = Configuration.objects.all()
    new_configurations = []
    updated_configurations = []
    for conf in response:
        exists = current_configurations.filter(converter_id=conf['id'])
        if exists.count() > 0:
            updated = exists[0]
            updated.name = conf['name']
            updated.configuration = conf['configuration']
            updated_configurations.append(updated)
        else:
            new_configurations.append(Configuration(converter_id=conf['id'], name=conf['name'],
                                                    configuration=conf['configuration']))
    Configuration.objects.bulk_update(updated_configurations, ['name', 'configuration'])
    if len(new_configurations):
        Configuration.objects.bulk_create(new_configurations)
    return


def avilon_custom_task():
    """
    Особая задача для Авилон Премиум Волгоградка. С прайса Авилон Seres Aito берёт по одному, самому дешёвому
    автомобилю, по фильтрам. Обновляет файл на ftp, который у Авилон Премиум Волгоградка прописан
    в поле Добавить к прайсу.
    :return:
    """
    avilon_seres_aito_task = ConverterTask.objects.get(pk=51)
    avilon_premium_volgogradka_task = ConverterTask.objects.get(pk=22)

    source_file = f'http://ph.onllline.ru/{avilon_seres_aito_task.price}'
    file_to_update = avilon_premium_volgogradka_task.add_to_price.replace('http://ph.onllline.ru/', '')
    temp_file = 'temp/avilon_custom_task.xlsx'

    # Обновляю Авилон SERES AITO
    get_price(task=avilon_seres_aito_task)

    # Сортирую по Марке, Модели и Цене
    df = pd.read_csv(source_file, decimal=',', sep=';', header=0, encoding='cp1251')
    df = df.sort_values(by=['Марка', 'Модель', 'Цена'])

    # Фильтрую и беру самый дешёвый по каждому фильтру
    filters = [
        {'Марка': 'AITO', 'Модель': 'M5'},
        {'Марка': 'Seres', 'Модель': 'Aito M7'}
    ]
    result_df = pd.DataFrame()
    for filter_cond in filters:
        filtered_df = df
        for key, value in filter_cond.items():
            filtered_df = filtered_df[filtered_df[key] == value]
        if not filtered_df.empty:
            result_df = pd.concat([result_df, filtered_df.iloc[[0]]], ignore_index=True)

    # Сохраняю в xlsx, отправляю на ftp
    result_df.T.reset_index().T.to_excel(temp_file, sheet_name='data', header=False, index=False)
    with open(temp_file, 'rb') as file:
        file_content = file.read()
        save_on_ftp(file_to_update, file_content)
