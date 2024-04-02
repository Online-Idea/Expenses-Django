import json

from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from libs.services.utils import split_daterange

# Цвета заливок
BLUE_FILL = PatternFill(start_color='DEE6EF', end_color='DEE6EF', fill_type='solid')
RED_FILL = PatternFill(start_color='E0C2CD', end_color='E0C2CD', fill_type='solid')
GREEN_FILL = PatternFill(start_color='DDE8CB', end_color='DDE8CB', fill_type='solid')
FONT = Font(name="Calibri", bold=True)


def get_srav_data(request, form, view) -> dict:
    """
    Возвращает данные формы. Для Выдачи и Сравнительной
    :return: словарь с данными формы
    """
    # Период
    daterange_str = form.cleaned_data.get('daterange')
    daterange = split_daterange(daterange_str)
    # Выбранные марки
    marks_checked = [m for m in request.POST.getlist('mark_checkbox')]
    # Выбранные регионы
    regions_checked = [r for r in request.POST.getlist('region_checkbox')]
    if view == 'comparison':
        # Дилер для сравнения
        dealer_for_comparison = form.cleaned_data.get('dealer') if 'dealer' in form.cleaned_data else ''

    # Общие
    filter_params = {
        'daterange': daterange_str,
    }

    # Отдельно для представления
    if view == 'parsed_ads':
        filter_params['mark_id__in'] = marks_checked
        filter_params['region__in'] = regions_checked
    elif view == 'comparison':
        filter_params['autoru_parsed_ad__mark_id__in'] = marks_checked
        filter_params['autoru_parsed_ad__region__in'] = regions_checked

    context = {
        'form': form,
        'marks_checked': json.dumps(marks_checked),
        'regions_checked': json.dumps(regions_checked),
        'datefrom': daterange['start'],
        'dateto': daterange['end'],
        'datefrom_dt': daterange['from'],
        'dateto_dt': daterange['to'],
        'filter_params': filter_params,
    }

    if view == 'comparison':
        context['dealer_for_comparison'] = dealer_for_comparison

    return context


def format_comparison(wb: Workbook, sheet_name: str, dealer_for_comparison: str = ''):
    """
    Форматирует эксель лист в читабельный вид
    :param wb:
    :param sheet_name:
    :param dealer_for_comparison:
    :return:
    """
    ws = wb[sheet_name]
    # Буквы нужных столбцов
    car_params = []
    price_cols = []
    difference_cols = []
    stock_cols = []
    position_actual_col = []

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col).value
        col_letter = get_column_letter(col)
        if cell in ['Марка', 'Модель', 'Комплектация', 'Модификация', 'Год']:
            car_params.append(col)
        elif cell == 'Имя дилера':
            dealer_col = col_letter
        elif 'цен' in cell.lower():
            price_cols.append(col_letter)
        elif cell == 'Ссылка':
            link_col = col_letter
        elif cell == 'В наличии' or cell == 'Под заказ':
            stock_cols.append(col_letter)
        elif 'Позиция' in cell:
            position_actual_col.append(col_letter)

        if 'Разница' in cell:
            difference_cols.append(col_letter)

    # Границы над уникальными автомобилями
    border = Border(top=Side(style='medium', color='000000'))
    prev_car, curr_car = '', ''
    for row in range(2, ws.max_row + 1):
        for car_param in car_params:
            curr_car += str(ws.cell(row=row, column=car_param).value)

        if curr_car != prev_car:
            for col2 in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col2)
                cell.border = border

        prev_car = curr_car
        curr_car = ''

    last_row = ws.max_row + 1
    for col2 in range(1, ws.max_column + 1):
        cell = ws.cell(row=last_row, column=col2)
        cell.border = border

    # Меняю ширину столбцов
    for column in ws.columns:
        ws[column[0].coordinate].alignment = Alignment(wrap_text=True)
        max_length = 0
        column_name = column[0].column_letter

        if column_name == link_col:
            ws.column_dimensions[column_name].width = 20
        elif column_name in price_cols or column_name in stock_cols or column_name in position_actual_col:
            ws.column_dimensions[column_name].width = 13
        else:
            for cell in column:
                if cell.row == 1:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_name].width = adjusted_width

    # Выделяю голубым нашего дилера
    for row in range(2, ws.max_row + 1):
        if ws[dealer_col + str(row)].value == dealer_for_comparison:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = BLUE_FILL

        # Выделяю красным тех кто дешевле нас и зеленым тех кто дороже нас
        for col in difference_cols:
            cell = ws[col + str(row)]
            if cell.value:
                if cell.value < 0:
                    cell.fill = RED_FILL
                elif cell.value > 0:
                    cell.fill = GREEN_FILL

    # Высота первой строки в 3 строки
    ws.row_dimensions[1].height = 45

    # Выравнивание по центру первой строки и жирный шрифт
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = FONT

    # Закрепляю строки и столбцы
    ws.freeze_panes = 'C2'

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    return wb
