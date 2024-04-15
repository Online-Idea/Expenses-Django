import os
import shutil
from datetime import datetime, timedelta
from io import BytesIO

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.protection import WorkbookProtection

from applications.autoconverter.converter import save_on_ftp
from libs.services.management.commands.online_idea_bot import online_idea_bot
from libs.services.utils import last_30_days, xlsx_column_width
from .models import *
from ..teleph.models import TelephCall


def export_calls_to_file() -> str:
    """
    Экспорт звонков в xlsx
    :return: имя файла
    """
    # РОЛЬФ клиенты которым это нужно
    # rolf_clients = {
    #     'Rolf_scoda_Center': 'РОЛЬФ Skoda Центр',
    #     'Rolf_Jetta_Center': 'РОЛЬФ Jetta Центр',
    #     'Rolf_moskvich_Himki': 'РОЛЬФ Москвич Химки',
    #     'Rolf_moskvich_Center': 'РОЛЬФ Москвич Центр',
    # }
    rolf_clients = {
        'porsche_nevsky': 'Порше Невский',
        'rolf_JLR_oct_spb': 'Рольф JLR Октябрьский СПБ',
        'rolf_multibrand_oct_spb': 'Рольф Мультибренд спб',
        'rolf_oktyabr_genesis_spb': 'Рольф Генезис СПБ',
        'rolf_viteb_exeed_spb': 'Рольф EXEED Витебский',
    }
    # Звонки из базы
    calls = TelephCall.objects.filter(client__teleph_id__in=rolf_clients.keys()) \
        .order_by('client_id', '-datetime').values()
    df = pd.DataFrame.from_records(calls)

    # Обработка столбцов
    # df['datetime'] = df['datetime'].dt.tz_convert(None)
    df['month_year'] = df['datetime'].dt.strftime('%m.%Y')
    df['datetime'] = df['datetime'].dt.strftime('%Y.%m.%d %H:%M:%S')
    df['moderation'] = df['moderation'].replace({'М': 'Авто.ру', 'М(Б)': 'Авто.ру', 'М(З)': 'Авто.ру',
                                                 'Доп.ресурсы': 'Дром', np.nan: 'Дром'})
    df['client_id'] = df['client_id'].replace(rolf_clients)

    # Оставляю нужные столбцы
    df = df[['client_id', 'datetime', 'num_from', 'mark', 'model', 'target', 'moderation', 'call_price', 'month_year']]

    # Переименовываю заголовки
    df = df.rename(columns={
        'client_id': 'Клиент',
        'datetime': 'Дата и время',
        'num_from': 'Номер клиента',
        'mark': 'Марка',
        'model': 'Модель',
        'target': 'Целевой',
        'moderation': 'Площадка',
        'call_price': 'Стоимость звонка',
        'month_year': 'Месяц, год',
    })

    # Открываю xlsx
    file_name = f'ROLF_stats_full.xlsx'
    file_path = f'temp/ROLF_stats/'
    book = load_workbook(f'{file_path}{file_name}')
    calls_sheet = book['Звонки']

    # Удаляю прошлые и вставляю новые
    calls_sheet.delete_rows(1, calls_sheet.max_row)
    for row in dataframe_to_rows(df, index=False, header=True):
        calls_sheet.append(row)

    custom_width = {
        'Площадка': 15,
        'Стоимость звонка': 15,
    }
    calls_sheet = xlsx_column_width(calls_sheet, custom_width)

    # # Меняю ширину столбцов
    # for column in calls_sheet.columns:
    #     max_length = 0
    #     column_letter = column[0].column_letter
    #     column_name = column[0].value
    #
    #     if column_name in ['Площадка', 'Стоимость звонка']:
    #         calls_sheet.column_dimensions[column_letter].width = 15
    #     else:
    #         for cell in column:
    #             if cell.row == 1:
    #                 continue
    #             try:
    #                 if len(str(cell.value)) > max_length:
    #                     max_length = len(str(cell.value))
    #             except:
    #                 pass
    #         adjusted_width = (max_length + 4)
    #         calls_sheet.column_dimensions[column_letter].width = adjusted_width

    # Сохраняю
    book.save(f'{file_path}{file_name}')
    save_on_ftp(f'{file_path}{file_name}')

    # Копия файла со скрытыми звонками и инструкцией
    file_nocalls = 'ROLF_stats.xlsx'
    os.remove(f'{file_path}{file_nocalls}')
    shutil.copy(f'{file_path}{file_name}', f'{file_path}{file_nocalls}')
    book_nocalls = load_workbook(f'{file_path}{file_nocalls}')

    book_nocalls['Звонки'].sheet_state = 'hidden'
    book_nocalls['Инструкция'].sheet_state = 'hidden'

    book_nocalls.security = WorkbookProtection()
    book_nocalls.security.workbookPassword = '8kjDF4ts#f9&'
    book_nocalls.security.lockStructure = True

    book_nocalls.save(f'{file_path}{file_nocalls}')
    save_on_ftp(f'{file_path}{file_nocalls}')

    # Сохраняю в xlsx
    # with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
    #     df.T.reset_index().T.to_excel(writer, sheet_name='Звонки', header=False, index=False)

    return file_name


def export_calls_for_callback(from_: str = None, to: str = None):
    """
    Экспорт звонков на перезвон
    :param from_:
    :param to:
    :return:
    """
    if not from_ or not to:
        from_, to = last_30_days()
    else:
        datetime_format = '%Y.%m.%d %H:%M:%S'
        from_ = datetime.strptime(from_, datetime_format)
        to = datetime.strptime(to, datetime_format)

    minus_3_days = to.date() - timedelta(days=3)
    client_filter = ['avilon_premium_legenda_new', 'avilon_premium_vol_new', 'Gate', 'avilon_bmw_new',
                     'avilon_exeed_new', 'avilon_foton_new', 'Avilon_rising_new', 'FnGroup_new']
    calls = TelephCall.objects.filter(client__teleph_id__in=client_filter, datetime__gte=from_, datetime__lte=to) \
        .values('client__name', 'datetime', 'num_from', 'target', 'call_status', 'comment')

    statuses = [
        'Сорвался',
        'Перевели, но Отдел продаж не взял трубку',
        'КЦ не взял трубку',
        'Не слышно клиента',
        'Обратный звонок. Клиент не взял трубку',
        'Никто не взял трубку',
    ]

    df = pd.DataFrame.from_records(calls)
    df['phone_from_comment'] = df['comment'].str.extract(r'\+(\d{11})')
    df['num_from'] = df.apply(
        lambda x: x['phone_from_comment'] if pd.notnull(x['phone_from_comment']) else x['num_from'], axis=1)
    df['date'] = df['datetime'].dt.date
    # Только мобильные. Мобильными считаю тех кто начинается на 79
    df = df[df['num_from'].astype(str).str.startswith('79')]

    calls_to_callback = {}
    unique_clients = df['client__name'].unique()
    for client in unique_clients:
        client_df = df[df['client__name'] == client]

        # Номера за прошедший день с нужными статусами
        statuses_df = client_df[(client_df['call_status'].str.contains('|'.join(statuses))) &
                                (client_df['date'] == to.date())]

        unique_phones = statuses_df['num_from'].unique()
        for phone in unique_phones:
            phone_df = client_df[client_df['num_from'] == phone]
            # Если по этому телефону не было целевых
            if phone_df[phone_df['target'].isin(['Да', 'ПМ - Целевой'])].empty:
                # Если по этому телефону не звонили последние 3 дня
                if phone_df[(phone_df['date'] >= minus_3_days) & (phone_df['date'] < to.date())].empty:
                    try:
                        phone = int(phone)
                    except ValueError:
                        phone = str(phone)
                    # Добавляем в прозвон
                    if client in calls_to_callback:
                        calls_to_callback[client] += [phone]
                    else:
                        calls_to_callback[client] = [phone]

    # Сохраняю в xlsx. Каждый клиент на отдельном листе
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    for key in calls_to_callback:
        sheet = wb.create_sheet(key)
        sheet.cell(row=1, column=1, value='Телефон')
        sheet.column_dimensions['A'].width = 15

        for i in range(len(calls_to_callback[key])):
            sheet.cell(row=i + 2, column=1, value=calls_to_callback[key][i])

    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    yesterday_fmt = to.strftime('%d.%m.%Y')
    # Этот chat.id из группы прозвона
    online_idea_bot.send_document('-1001839691903', (f'Перезвон {yesterday_fmt}.xlsx', virtual_workbook))
    # online_idea_bot.send_document('289346624', (f'Перезвон {yesterday_fmt}.xlsx', virtual_workbook))  # Это мой чат с ботом
    # online_idea_bot.send_message('-1001839691903', 'test')

    return wb
