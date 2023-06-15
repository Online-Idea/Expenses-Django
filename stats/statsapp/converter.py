import ftplib
from ftplib import FTP
from functools import reduce
from pathlib import Path
import os

import pandas.core.series
import requests
import datetime
import xml.etree.ElementTree as ET
from openpyxl import Workbook
# import xlsxwriter
import pandas as pd
from pandas import DataFrame
from telebot.types import InputFile

from stats.settings import env
from statsapp.models import *
from statsapp.management.commands.bot import bot


# Список Конфигураций: POST http://151.248.118.19/Api/Configurations/GetList
# Список Папок с фото: POST http://151.248.118.19/Api/Stock/GetClients

# Прогон шаблона (3 запроса)
# POST http://151.248.118.19/Api/Stock/StartProcess
# POST http://151.248.118.19/Api/Stock/GetProcessStep
# POST http://151.248.118.19/Api/Log/GetByProcessId

def get_converter_tasks():
    active_tasks = ConverterTask.objects.filter(active=True)
    return active_tasks


def get_price(task):
    """
    Полный цикл для одного клиента.
    :param task: строка из таблицы Задачи конвертера
    """
    template = converter_template(task)
    client_slug = task.client.slug
    client_name = task.client.name
    process_id = converter_post(task)
    price = converter_process_result(process_id, client_slug, template, task)
    logs = converter_logs(process_id)
    logs_xlsx = logs_to_xlsx(logs, template, client_slug)
    bot_messages(logs, logs_xlsx, price, client_slug, client_name)
    save_on_ftp(logs_xlsx)
    os.remove(logs_xlsx)
    print(f'Клиент {client_slug} - прайс готов')
    return


def converter_template(task):
    """
    Из стока клиента делает шаблон для конвертера
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    # Сохраняю сток клиента, делаю по нему шаблон для конвертера
    slug = task.client.slug
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    stock_path = f'converter/{slug}/stocks/stock_{slug}_{file_date}'

    # Получаю сток
    if task.stock_source == 'Ссылка':
        response = requests.get(url=task.stock_url)
    elif task.stock_source == 'POST-запрос':
        data = {
            'login': task.stock_post_login,
            'password': task.stock_post_password,
        }
        response = requests.post(url=task.stock_post_host, data=data)

    # Получаю тип файла стока
    content_type = response.headers['content-type']
    if 'text/xml' in content_type or 'application/xml' in content_type:
        stock_path += '.xml'
        content_type = 'xml'
    elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
        stock_path += '.xlsx'
        content_type = 'xlsx'
    elif 'text/csv' in content_type:
        stock_path += '.csv'
        content_type = 'csv'

    # Сохраняю сток на ftp
    os.makedirs(os.path.dirname(stock_path), mode=0o755, exist_ok=True)
    with open(stock_path, mode='wb') as file:
        file.write(response.content)
    # with open(stock_path, mode='w', encoding=task.stock_fields.encoding) as file:
    #     file.write(response.text)
    save_on_ftp(stock_path)

    # Путь шаблона
    template_path = f'converter/{slug}/templates/template_{slug}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    task.template = template_path
    task.save()

    # Разные функции в зависимости от типа файла стока
    if content_type == 'xml':
        template = template_xml(stock_path, template_path, task)
    elif content_type == 'xlsx':
        template = template_xlsx_or_csv(stock_path, 'xlsx', template_path, task)
    elif content_type == 'csv':
        template = template_xlsx_or_csv(stock_path, 'csv', template_path, task)
    else:
        return 'Неверный формат файла, должен быть xml или xlsx'

    save_on_ftp(template_path)
    os.remove(stock_path)

    return template


def template_xml(stock_path, template_path, task):
    """
    Шаблон из xml стока
    :param stock_path: путь к файлу стока
    :param template_path: путь к файлу шаблона
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    # XML tree
    tree = ET.parse(stock_path)
    root = tree.getroot()

    # XLSX шаблон
    workbook = Workbook()
    workbook.active.title = 'Шаблон'
    sheet = workbook.active

    # Заголовки шаблона
    template_col = StockFields.TEMPLATE_COL
    for k, col in template_col.items():
        sheet.cell(row=1, column=col[1] + 1, value=col[0])

    # Данные шаблона
    fields = task.stock_fields
    exception_col = ['modification_code', 'options_code', 'images', 'modification_explained']
    for i, car in enumerate(root.iter(fields.car_tag)):
        if stock_xml_filter(car, task):
            # Обычные поля
            for field in fields._meta.fields:
                field_val = getattr(fields, field.name)
                # Если не пусто И поле в полях шаблона И поле НЕ в исключениях
                if field_val and field.name in template_col and field.name not in exception_col:
                    cell = car.findtext(field_val)
                    try:
                        if cell.isnumeric():
                            cell = int(cell)
                    except AttributeError:
                        pass
                    sheet.cell(row=i + 2, column=template_col[field.name][1] + 1, value=cell)

            # Поля-исключения
            if ',' in fields.modification_code:  # Код модификации
                # Разделяет по запятой в список если есть значение. Убирает запятую из данных стока
                mod = [car.findtext(f).replace(',', '') for f in fields.modification_code.split(', ') if
                       car.findtext(f)]
                sheet.cell(row=i + 2, column=template_col['modification_code'][1] + 1, value=' | '.join(mod))
            else:
                sheet.cell(row=i + 2, column=template_col['modification_code'][1] + 1,
                           value=car.findtext(fields.modification_code))

            if fields.options_code:
                options = multi_tags(fields.options_code, car)  # Опции
                sheet.cell(row=i + 2, column=template_col['options_code'][1] + 1, value=options)

            if fields.images:
                if 'авилон' in task.client.name.lower():
                    # Особая обработка фото от Авилона
                    images = avilon_photos(fields.images, car)
                else:
                    images = multi_tags(fields.images, car)  # Фото клиента
                sheet.cell(row=i + 2, column=template_col['images'][1] + 1, value=images)

            if ',' in fields.modification_explained:  # Расш. модификации
                mod = [car.findtext(f) for f in fields.modification_explained.split(', ') if car.findtext(f)]
                sheet.cell(row=i + 2, column=template_col['modification_explained'][1] + 1, value=' | '.join(mod))
            else:
                sheet.cell(row=i + 2, column=template_col['modification_explained'][1] + 1,
                           value=car.findtext(fields.modification_explained))

            # Для обработки прайса когда нужно смотреть по стоку. Добавляю столбец к шаблону
            extras = ConverterExtraProcessing.objects.filter(converter_task=task, source='Сток')
            if extras:
                for extra in extras:
                    conditionals = Conditionals.objects.filter(converter_extra_processing=extra.id)
                    for cond in conditionals:
                        column_name = cond.field
                        value = multi_tags(cond.field, car)

                        if column_name not in template_col:
                            max_column = len(template_col)
                            template_col[column_name] = (column_name, max_column)
                            sheet.cell(row=1, column=max_column + 1, value=column_name)

                        column_number = template_col[column_name][1] + 1

                        sheet.cell(row=i + 2, column=column_number, value=value)

    # Удаляю пустые строки
    for row in reversed(range(1, sheet.max_row + 1)):
        if all(cell.value is None for cell in sheet[row]):
            sheet.delete_rows(row)

    workbook.save(filename=template_path)
    return pd.read_excel(template_path, decimal=',')


def avilon_photos(field: str, element: ET.Element) -> str:
    """
    Особая настройка для Авилона. Если брать их фото в том порядке как они идут в xml,
    то порядок фото получается случайным. А так как Авилон к каждому фото указывает угол,
    с которого они снимали, то можно по этому углу выставлять в красивый порядок фото
    :param field: имя тэга с фото
    :param element: элемент автомобиля
    :return: фото в одной строке, разделённые запятой
    """

    def get_by_view_type(tags, view_type):
        return [tag for tag in tags if tag.attrib['view_type'] == view_type]

    tags = element.findall(field)

    exterior = get_by_view_type(tags, 'exterior')
    # Первое фото немного боком влево
    for element in exterior:
        if element.attrib['angle'] in ['40', '320']:
            element.set('angle', '-40')
            break
    exterior.sort(key=lambda x: float(x.get('angle')))

    detail = get_by_view_type(tags, 'detail')

    interior = get_by_view_type(tags, 'interior')

    images = [image.text for image in exterior]
    images += [image.text for image in detail]
    images += [image.text for image in interior]
    return ', '.join(images)


def stock_xml_filter(car, task):
    """
    Проверяет автомобиль из xml стока по фильтрам из ConverterFilters
    :param car: node автомобиля из xml стока
    :param task: task (запись) из таблицы Задачи конвертера
    :return: True если фильтры пройдены
    """
    filters = ConverterFilters.objects.filter(converter_task=task)
    dict_filters, stock_fields, result = [], [], []

    # Перевожу фильтры в словарь вида: {'values': values, 'condition': condition, 'field': field}
    for f in filters:
        # Значения
        if '`' in f.value:
            values = [val.replace('`', '').strip() for val in f.value.split('`,')]
        else:
            values = [f.value]

        # Поля (теги)
        if '/' in f.field:  # Путь к детям
            parent = f.field.split('/')[0]

            if '@' not in f.field:  # Дети
                stock_fields = [tag.text for tag in car.find(parent)]

            else:  # Атрибуты
                attribute_name = f.field.split('@')[1].split('=')[0]
                attribute_value = f.field.split('"')[1].replace('"', '')
                for tag in car.find(parent):
                    for tag_attribute in tag.attrib.items():
                        if tag_attribute == (attribute_name, attribute_value):
                            stock_fields = [tag.text]

        else:  # Просто один тег
            stock_fields = [car.findtext(f.field)]

        for field in stock_fields:
            dict_filters.append({
                'values': values,
                'condition': f.condition,
                'field': field,
            })

    for sf in dict_filters:
        for value in sf['values']:
            result.append(xml_filter_conditions(value, sf['condition'], sf['field']))

    # Если длина фильтров равна сумме результатов значит каждый фильтр вернул True, соответственно автомобиль подходит
    return len(dict_filters) == sum(result)


def xml_filter_conditions(value, condition, stock_field):
    """
    Условия фильтра
    :param value: значение для проверки
    :param condition: условие
    :param stock_field: значение из стока
    :return: True если условие выполнено
    """
    if condition == ConverterFilters.GREATER_THAN:
        if not stock_field:
            stock_field = 0
        return eval(f'{int(stock_field)} {condition} {int(value)}')
    elif condition == ConverterFilters.LESS_THAN:
        if not stock_field:
            stock_field = 0
        return eval(f'{int(stock_field)} {condition} {int(value)}')
    elif 'with' not in condition:
        return eval(f'"{value}" {condition} "{stock_field}"')
    elif condition == ConverterFilters.STARTS_WITH:
        return stock_field.startswith(value)
    elif condition == ConverterFilters.NOT_STARTS_WITH:
        return not (stock_field.startswith(value))
    elif condition == ConverterFilters.ENDS_WITH:
        return stock_field.endswith(value)
    elif condition == ConverterFilters.NOT_ENDS_WITH:
        return not (stock_field.endswith(value))


def template_xlsx_or_csv(stock_path, filetype, template_path, task):
    """
    Шаблон из xlsx или csv стока
    :param stock_path: путь к файлу стока
    :param filetype: тип файла: 'xlsx' или 'csv'
    :param template_path: путь к файлу шаблона
    :param task: строка из таблицы Задачи конвертера
    :return: шаблон как pandas dataframe
    """
    if filetype == 'xlsx':
        df_stock = pd.read_excel(stock_path, decimal=',')
    elif filetype == 'csv':
        df_stock = pd.read_csv(stock_path, decimal=',', sep=';', header=0, encoding='cp1251')
    else:
        return 'Неверный формат файла, должен быть xlsx или csv'

    # Проверяю если сток это наш прайс. На случай если прайс готов и нужно только фото подставить
    # Если первые 4 столбца совпадают с our_price_first_4_cols И Исходный VIN в столбцах
    our_price_first_4_cols = ['Марка', 'Модель', 'Комплектация', 'Авто.ру Комплектация']
    if list(df_stock.columns[:4]) == our_price_first_4_cols and 'Исходный VIN' in df_stock.columns:
        df_stock.loc[df_stock['Фото'].str.contains('gallery'), 'Фото'] = ''
        df_stock.T.reset_index().T.to_excel(template_path, sheet_name='Шаблон', header=False, index=False)
        return df_stock

    df_stock = stock_xlsx_filter(df_stock, task)

    fields = StockFields.objects.filter(pk=task.stock_fields.id)
    fields = fields.values()[0]
    template_col = StockFields.TEMPLATE_COL

    # Меняю имена столбцов
    swapped_cols = {v: template_col[k][0] for k, v in fields.items() if k in template_col}
    df_stock.rename(columns=swapped_cols, inplace=True)

    # Добавляю недостающие столбцы
    cur_cols = list(df_stock.columns.values)
    for k, col in template_col.items():
        if col[0] not in cur_cols:
            df_stock[col[0]] = ''

    # Для обработки прайса когда нужно смотреть по стоку. Добавляю столбец к шаблону
    extras = ConverterExtraProcessing.objects.filter(converter_task=task, source='Сток')
    if extras:
        for extra in extras:
            conditionals = Conditionals.objects.filter(converter_extra_processing=extra.id)
            for cond in conditionals:
                column_name = cond.field
                if column_name not in template_col:
                    max_column = len(template_col)
                    template_col[column_name] = (column_name, max_column)

    # Оставляю только нужные столбцы в том порядке как в template_col
    df_stock = df_stock[[v[0] for k, v in template_col.items()]]
    # В Опции и пакеты заменяю переносы строк на пробел
    df_stock['Опции и пакеты'].replace(r'\n', ' ', regex=True, inplace=True)

    df_stock.T.reset_index().T.to_excel(template_path, sheet_name='Шаблон', header=False, index=False)

    return df_stock


def stock_xlsx_filter(df, task):
    """
    Проверяет автомобиль из xlsx стока по фильтрам из ConverterFilters
    :param df: xlsx сток в виде pandas dataframe
    :param task: task (запись) из таблицы Задачи конвертера
    :return: Отфильтрованный dataframe
    """
    filters = ConverterFilters.objects.filter(converter_task=task)
    filter_strings = []
    for f in filters:
        filter_or = []
        if '`' in f.value:
            values = [val.replace('`', '').strip() for val in f.value.split('`,')]
        else:
            values = [f.value]

        for value in values:
            if f.condition == ConverterFilters.CONTAINS:
                filter_or.append(f'(df["{f.field}"].str.contains({value})')
            elif f.condition == ConverterFilters.NOT_CONTAINS:
                filter_or.append(f'(~df["{f.field}"].str.contains({value})')
            elif f.condition == ConverterFilters.EQUALS:
                filter_or.append(f'(df["{f.field}"] == "{value}")')
            elif f.condition == ConverterFilters.NOT_EQUALS:
                filter_or.append(f'(~df["{f.field}"] == "{value}")')
            elif f.condition == ConverterFilters.GREATER_THAN:
                filter_or.append(f'(df["{f.field}"] > "{value}")')
            elif f.condition == ConverterFilters.LESS_THAN:
                filter_or.append(f'(df["{f.field}"] < "{value}")')
            elif f.condition == ConverterFilters.STARTS_WITH:
                filter_or.append(f'(df["{f.field}"].str.startswith("{value}"))')
            elif f.condition == ConverterFilters.NOT_STARTS_WITH:
                filter_or.append(f'~(df["{f.field}"].str.startswith("{value}"))')
            elif f.condition == ConverterFilters.ENDS_WITH:
                filter_or.append(f'(df["{f.field}"].str.endswith("{value}"))')
            elif f.condition == ConverterFilters.NOT_ENDS_WITH:
                filter_or.append(f'~(df["{f.field}"].str.endswith("{value}"))')

        filter_strings.append(f'({" | ".join(filter_or)})')

    return eval(f'df.loc[{" & ".join(filter_strings)}]') if filter_strings else df


def price_extra_processing(df: DataFrame, task: ConverterTask, template: DataFrame) -> DataFrame:
    """
    Меняет данные в прайсе по условию
    :param df: прайс в виде pandas dataframe
    :param task: task (запись) из таблицы Задачи конвертера
    :param template: шаблон в виде pandas dataframe
    :return: Готовый прайс с изменениями
    """
    changes = ConverterExtraProcessing.objects.filter(converter_task=task)

    # Если изменение зависит от данных в стоке то сначала эти данные добавляются к шаблону.
    # Потом шаблон передаётся сюда и добавляется к прайсу.
    if changes.filter(source='Сток'):
        template = template.add_suffix('_template')
        df = pd.merge(df, template, left_index=True, right_index=True)

    # Для каждого изменения по клиенту
    for change in changes:
        conditionals = Conditionals.objects.filter(converter_extra_processing=change.id)
        masks = []

        # Для каждого условия в изменении
        for cond in conditionals:

            if change.source == 'Сток':
                cond.field += '_template'

            # Если несколько значений
            if '`' in cond.value:
                values = [val.replace('`', '').strip() for val in cond.value.split('`,')]
            else:
                values = [cond.value]

            # Для каждого значения
            or_masks = []
            for value in values:
                if value.isdigit():
                    value = int(value)

                if cond.condition == ConverterFilters.CONTAINS:
                    or_masks.append(df[cond.field].str.contains(value))
                elif cond.condition == ConverterFilters.NOT_CONTAINS:
                    or_masks.append(~df[cond.field].str.contains(value))
                elif cond.condition == ConverterFilters.EQUALS:
                    or_masks.append(df[cond.field] == value)
                elif cond.condition == ConverterFilters.NOT_EQUALS:
                    or_masks.append(df[cond.field] != value)
                elif cond.condition == ConverterFilters.GREATER_THAN:
                    or_masks.append(df[cond.field] > value)
                elif cond.condition == ConverterFilters.LESS_THAN:
                    or_masks.append(df[cond.field] < value)
                elif cond.condition == ConverterFilters.STARTS_WITH:
                    or_masks.append(df[cond.field].str.startswith(value))
                elif cond.condition == ConverterFilters.NOT_STARTS_WITH:
                    or_masks.append(~df[cond.field].str.startswith(value))
                elif cond.condition == ConverterFilters.ENDS_WITH:
                    or_masks.append(df[cond.field].str.endswith(value))
                elif cond.condition == ConverterFilters.NOT_ENDS_WITH:
                    or_masks.append(~df[cond.field].str.endswith(value))

            # Несколько значений через ИЛИ
            combined_or_mask = reduce(lambda x, y: x | y, or_masks)
            combined_or_mask = combined_or_mask.fillna(False)
            masks.append(combined_or_mask)

        # Объединяю маски
        combined_mask = reduce(lambda x, y: x & y, masks)
        # Проставляю новые значения
        df.loc[combined_mask, change.price_column_to_change] = change.new_value

    df = df.drop(df.filter(regex='_template').columns, axis=1)

    return df


def multi_tags(field, element):
    """
    Обрабатывает поля для которых данные собираются из нескольких тегов
    :param field: поле
    :param element: элемент из xml
    :return: готовые данные для шаблона
    """
    result = []

    if '/' in field:
        parent = field.split('/')[0]
        parent_element = element.find(parent)
        if parent_element:
            if '@' not in field:  # Если тег с детьми и нужно значение детей
                result = [tag.text for tag in parent_element]
            else:  # Если тег с детьми и из детей нужен атрибут
                attribute = field.split('@')[1]
                result = [tag.attrib[attribute] for tag in parent_element]
    else:
        if '@' not in field:  # Если тег несколько раз повторяется и нужно значение
            tags = element.findall(field)
            if tags:
                result = [tag.text for tag in tags]
        else:  # Если тег несколько раз повторяется и из него нужен атрибут
            tag_name, attribute = field.split('@')
            tags = element.findall(tag_name)
            if tags:
                result = [tag.attrib[attribute] for tag in tags]

    return ' '.join(result)


def converter_post(task):
    """
    Отправляет шаблон и настройки по которым нужно прогнать конвертер
    :param task: task (запись) из таблицы Задачи конвертера
    :return: прайс как pandas dataframe
    """
    # Отправляю шаблон
    url = 'http://151.248.118.19/Api/Stock/StartProcess'

    configuration = task.configuration.configuration if task.configuration is not None else Configuration.DEFAULT
    payload = {
        'client': task.photos_folder.folder,
        'configuration': configuration,
        'frontLength': task.front,
        'backLength': task.back,
        'interiorsLength': task.interior,
        'onlySalon': task.salon_only,
    }
    template = open(task.template, 'rb')
    files = {'file': ('template.xlsx', template, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                      {'Expires': '0'})}
    response = requests.post(url=url, data=payload, files=files)
    template.close()
    os.remove(task.template)
    return response.json()['processId']


def converter_process_result(process_id, client, template, task):
    """
    Возвращает готовый прайс
    :param process_id: из converter_post
    :param client: имя клиента как slug - используется как имя папки клиента куда сохраняется прайс
    :param template: шаблон как pandas dataframe - если из шаблона нужны данные для прайса
    :param task: строка из таблицы Задачи конвертера
    """
    # Получаю прайс от конвертера по process_id
    url = 'http://151.248.118.19/Api/Stock/GetProcessResult'
    payload = {'processId': (None, process_id)}
    response = requests.post(url=url, files=payload)

    # Сохраняю прайс в xlsx
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    save_path_date = f'converter/{client}/prices/price_{client}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(save_path_date), exist_ok=True)
    with open(save_path_date, 'wb') as file:
        file.write(response.content)
    save_on_ftp(save_path_date)

    # Обработки прайса
    read_file = pd.read_excel(save_path_date, decimal=',')
    # Переношу Описание из шаблона если оно есть в шаблоне
    if 'Описание' in template.columns:
        read_file['Описание2'] = template['Описание']
        read_file.loc[read_file['Описание2'].notnull(), 'Описание'] = read_file['Описание2']
        read_file.drop(columns='Описание2', inplace=True)

    read_file['Описание'] = read_file['Описание'].replace('_x000d_', '', regex=True)

    # Убираю автомобили которые не расшифрованы (пустые столбцы Марка, Цвет либо Фото)
    read_file = read_file[(~read_file['Марка'].isnull()) &
                          (~read_file['Цвет'].isnull()) &
                          (~read_file['Фото'].isnull())]
    # Если Максималка пустая то считать её как сумму других скидок
    read_file.loc[read_file['Максималка'].isna(), 'Максималка'] = read_file[['Трейд-ин', 'Кредит', 'Страховка']].sum(
        axis=1)

    # Различные изменения прайса по условию
    read_file = price_extra_processing(read_file, task, template)

    read_file.fillna('', inplace=True)
    read_file = read_file.astype(str).replace(r'\.0$', '', regex=True)
    read_file = read_file.astype(str).replace('é', 'e', regex=True)

    # Сохраняю в csv
    save_path = f'converter/{client}/prices/price_{client}.csv'
    read_file.to_csv(save_path, sep=';', header=True, encoding='cp1251', index=False, decimal=',')
    save_on_ftp(save_path)

    os.remove(save_path_date)
    os.remove(save_path)
    return read_file


def converter_logs(process_id):
    """
    Логи конвертера
    :return: логи в виде текста
    """
    # Логи которые присылает конвертер
    url = 'http://151.248.118.19/Api/Log/GetByProcessId'
    payload = {'processId': process_id}
    response = requests.post(url=url, json=payload)
    logs = response.json()['log']
    logs = logs.replace(' ,', '')  # Убираю лишние запятые
    return logs


def logs_to_xlsx(logs, template, client):
    """
    Логи конвертера вместе с расшифровкой в xlsx
    :param logs: логи от converter_logs
    :param template: шаблон от converter_template
    :param client: клиент как client_slug для имени папки и файла
    :return: xlsx файл логов
    """
    lookup_cols = {
        # База из лога: (Имя столбца с кодом, Имя столбца с расшифровкой)
        'Модификации': ('Код модификации', 'Расш. модификации'),
        'Цвета': ('Код цвета', 'Расш. цвета'),
        'Интерьеры': ('Код интерьера', 'Расш. интерьера'),
        'Комплектации': ('Код модификации', 'Расш. модификации'),
        # 'Опции': в логи идут только коды, без расшифровки
        # 'Фото': только количество без фото
    }

    # Переделываю логи в словарь
    lines = logs.split('\n')[:-2]  # Последние 2 убираю т.к. там Время обработки и пустая строка
    logs_dict = {}
    for line in lines:
        try:
            key = line.split('"')[1]
        except IndexError:
            continue
        start = line.index(':') + 1
        end = line.index(';')
        value = []
        for v in line[start:end].split(','):
            v = v.strip()
            if v.isnumeric():
                v = int(v)
            value.append(v)

        if key in ['Опции', 'Фото']:  # Опции без расшифровки, Фото только количество
            logs_dict[key] = pd.Series(value, name=key)
        else:  # Остальные это pandas dataframe в виде Кода и Расшифровки
            df2 = pd.Series(value, name='Код', dtype='string')
            joined = pd.merge(template, df2, left_on=lookup_cols[key][0], right_on='Код')
            joined.drop_duplicates(subset=[lookup_cols[key][0]], inplace=True)
            joined = joined[[lookup_cols[key][0], lookup_cols[key][1]]]
            logs_dict[key] = joined

    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    logs_save_path = f'converter/{client}/logs/log_{client}_{file_date}.xlsx'
    os.makedirs(os.path.dirname(logs_save_path), exist_ok=True)

    # Готовые логи в xlsx
    with pd.ExcelWriter(logs_save_path) as writer:
        for key, value in logs_dict.items():
            df = pd.DataFrame(value)
            # Такой длинный вариант чтобы убрать форматирование заголовков которое pandas применяет по умолчанию
            df.T.reset_index().T.to_excel(writer, sheet_name=key, header=False, index=False)
    return logs_save_path


def bot_messages(logs, logs_xlsx, price, client_slug, client_name):
    """
    Сообщения для телеграм бота
    :param logs: логи от converter_logs
    :param logs_xlsx: логи в xlsx от logs_to_xlsx
    :param price: прайс от converter_process_result
    :param client_slug: клиент как client_slug для имени папки и файла
    :param client_name: клиент как client_name для сообщения бота
    """
    # Прайс в csv
    file_date = str(datetime.datetime.now()).replace(' ', '_').replace(':', '-')
    price_save_path = f'converter/{client_slug}/prices/price_{client_slug}_{file_date}.csv'
    price.to_csv(price_save_path, sep=';', header=True, encoding='cp1251', index=False, decimal=',')

    # Отправка логов и прайса через бота телеграма
    chat_ids = ConverterLogsBotData.objects.all()
    logs = f'🟢 {client_name}\n\n{logs}'
    for chat_id in chat_ids:
        if len(logs) > 4095:  # У телеграма ограничение на 4096 символов в сообщении
            for x in range(0, len(logs), 4095):
                bot.send_message(chat_id.chat_id, logs[x:x + 4095])
        else:
            bot.send_message(chat_id.chat_id, logs)
        bot.send_document(chat_id.chat_id, InputFile(logs_xlsx))
        bot.send_document(chat_id.chat_id, InputFile(price_save_path))

    os.remove(price_save_path)
    return


def save_on_ftp(save_path):
    """
    Сохраняет файл на ftp
    :param save_path: полный путь к файлу
    """
    file_path = Path(save_path)
    with FTP('ph.onllline.ru', env('FTP_LOGIN'), env('FTP_PASSWORD')) as ftp, open(save_path, 'rb') as file:
        cd_tree(ftp, str(file_path.parents[0]))
        ftp.storbinary(f'STOR {file_path.name}', file)
        ftp.cwd('/')
    return


def cd_tree(ftp, path):
    """
    Создаёт папки на ftp если они не существуют
    :param ftp: FTP класс из ftplib
    :param path: путь который нужен на ftp
    """
    path = path.replace('\\', '/')  # Костыль для windows
    for folder in path.split('/'):
        try:
            ftp.cwd(folder)
        except ftplib.error_perm:
            ftp.mkd(folder)
            ftp.cwd(folder)
    return


def get_photo_folders():
    """ Получает от конвертера список папок с фото и добавляет новые в базу """
    url = 'http://151.248.118.19/Api/Stock/GetClients'
    response = requests.post(url).json()
    current_folders = [f.folder for f in PhotoFolder.objects.all()]
    new_folders = []
    for folder in response:
        if folder not in current_folders:
            new_folders.append(PhotoFolder(folder=folder))
    PhotoFolder.objects.bulk_create(new_folders)
    return


def get_configurations():
    """ Получает от конвертера конфигурации и добавляет/обновляет в базе """
    url = 'http://151.248.118.19/Api/Configurations/GetList'
    response = requests.post(url).json()
    current_configurations = Configuration.objects.all()
    new_configurations = []
    updated_configurations = []
    for conf in response:
        exists = current_configurations.filter(converter_id=conf['id'])
        if exists.count() > 0:
            updated = exists[0]
            updated.name = conf['name']
            updated.configuration = conf['configuration']
            updated_configurations.append(updated)
        else:
            new_configurations.append(Configuration(converter_id=conf['id'], name=conf['name'],
                                                    configuration=conf['configuration']))
    Configuration.objects.bulk_update(updated_configurations, ['name', 'configuration'])
    if len(new_configurations):
        Configuration.objects.bulk_create(new_configurations)
    return
